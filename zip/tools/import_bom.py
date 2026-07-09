#!/usr/bin/env python3
"""Импорт инженерного BOM перфоратора (разузловка + статус разработки) в базу.

Источник: zip/tools/sources/bom_cop3060mux.xlsx (снимок таблицы инженеров).
Выход: zip/data/bom.json — {machines:[{name, parts:[...], kits:[...], stats}]}.
Отдельная вкладка «Разработка» показывает воронку локализации по деталям
(Замер / РКД / ОПИ), ответственных, кол-во на машину, поставщиков, цены.

Мульти-машинный формат: файлы zip/tools/sources/bom_*.xlsx все подхватываются.
"""
import json
import re
from pathlib import Path
from datetime import date

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
SRC = ROOT / "tools" / "sources"


def s(v):
    return "" if v is None else str(v).strip()


def rows_of(ws):
    return [list(r) for r in ws.iter_rows(values_only=True)]


def parse_perforator(ws):
    rows = rows_of(ws)
    data = [r for r in rows[2:] if any(s(c) for c in r)]
    parts = []
    for r in data:
        g = lambda i: s(r[i]) if i < len(r) else ""
        epn, kv, desc = g(3), g(4), g(6)
        if not (epn or kv or desc):
            continue
        # уровень в разузловке
        lvl = 1 if g(0) else 2 if g(1) else 3 if g(2) else 0
        scheme = g(0) or g(1) or g(2)
        parts.append({
            "scheme": scheme, "level": lvl,
            "epiroc_pn": epn, "kv_pn": kv, "qty": g(5), "desc": desc,
            "owner": g(7),
            "st_izmer": g(8), "st_rkd": g(9), "st_opi": g(10),
            "suppliers": [x for x in (g(11), g(12), g(13)) if x],
            "price_comp": [x for x in (g(14), g(15), g(16)) if x],
            "price_our": g(17),
        })
    return parts


def parse_kits(ws):
    rows = rows_of(ws)
    data = [r for r in rows[1:] if any(s(c) for c in r)]
    kits = []
    for r in data:
        g = lambda i: s(r[i]) if i < len(r) else ""
        name, oem, kv = g(3), g(4), g(5)
        if not (name or oem or kv):
            continue
        kits.append({
            "kit_no": g(0), "pp": g(1), "scheme": g(2),
            "name": name, "oem_pn": oem, "kv_pn": kv, "qty": g(6),
        })
    return kits


def funnel(parts):
    """Сводка по этапам разработки."""
    def cnt(key):
        c = {}
        for p in parts:
            v = p.get(key) or "—"
            c[v] = c.get(v, 0) + 1
        return c
    with_status = sum(1 for p in parts if p["st_izmer"] or p["st_rkd"] or p["st_opi"])
    owners = {}
    for p in parts:
        if p["owner"]:
            owners[p["owner"]] = owners.get(p["owner"], 0) + 1
    return {
        "parts": len(parts), "with_status": with_status,
        "izmer": cnt("st_izmer"), "rkd": cnt("st_rkd"), "opi": cnt("st_opi"),
        "owners": sorted(owners.items(), key=lambda a: -a[1]),
        "with_kv_pn": sum(1 for p in parts if p["kv_pn"]),
    }


def machine_name(wb, fallback):
    for t in wb.sheetnames:
        m = re.search(r"COP\s*\w+", t)
        if m:
            return m.group(0).replace("  ", " ")
    return fallback


def main():
    machines = []
    for xlsx in sorted(SRC.glob("bom_*.xlsx")):
        wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        perf_ws = next((wb[t] for t in wb.sheetnames if "perforator" in t.lower()), None)
        kit_ws = next((wb[t] for t in wb.sheetnames if "kit" in t.lower()), None)
        parts = parse_perforator(perf_ws) if perf_ws else []
        kits = parse_kits(kit_ws) if kit_ws else []
        machines.append({
            "name": machine_name(wb, xlsx.stem), "file": xlsx.name,
            "parts": parts, "kits": kits, "stats": funnel(parts),
        })
        print(f"[{xlsx.name}] деталей {len(parts)}, наборов {len(kits)}, "
              f"со статусом {funnel(parts)['with_status']}")
    out = {"updated": date.today().isoformat(), "machines": machines}
    (DATA / "bom.json").write_text(json.dumps(out, ensure_ascii=False, indent=1),
                                   encoding="utf-8")
    print(f"→ zip/data/bom.json ({len(machines)} машин)")


if __name__ == "__main__":
    main()
