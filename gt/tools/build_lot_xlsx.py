#!/usr/bin/env python3
"""Два файла-деливерабла по лоту ЛУКОЙЛ:
1) lot_ocenka.xlsx — оценка каждой позиции (вилка USD, уверенность, основа, топ-поставщики)
2) lot_postavshiki.xlsx — поставщики/изготовители, сгруппированные по направлениям лота

Читает: gt/data/rfq_demand.json, gt/data/rfq_prices.json (накопленные ценовые волны),
подбор поставщиков — те же правила, что в build_rfq.py.
"""
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "tools"))
from build_rfq import CAT_KEYS, build_pool, dir_of, model_match, probability  # noqa: E402

D = lambda f: json.load(open(ROOT / "data" / f))

TH = Font(name="Arial", bold=True, size=10, color="FFFFFF")
TD = Font(name="Arial", size=10)
HFILL = PatternFill("solid", fgColor="1F3B57")
BORD = Border(*[Side(style="thin", color="D9D9D9")] * 4)
CONF_FILL = {"A": PatternFill("solid", fgColor="D9F2D9"),
             "B": PatternFill("solid", fgColor="FDEECD"),
             "C": PatternFill("solid", fgColor="F2F2F0")}


def sheet_header(ws, cols):
    ws.append([c[0] for c in cols])
    for i, (_, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        cell = ws.cell(row=ws.max_row, column=i)
        cell.font, cell.fill, cell.border = TH, HFILL, BORD
    ws.freeze_panes = "A2"


def main():
    demand = D("rfq_demand.json")["rows"]
    try:
        prices = {p["pn"].strip().upper(): p for p in D("rfq_prices.json")["prices"] if p.get("pn")}
    except FileNotFoundError:
        prices = {}
    pool = build_pool()

    # топ-поставщики на (направление, категория)
    assign_cache = {}

    def top_sups(direction, cat, n=3):
        key = (direction, cat)
        if key not in assign_cache:
            cands = []
            for sup in pool.values():
                if not (model_match(sup, direction) or re.search(CAT_KEYS.get(cat, "$^"), sup["what"], re.I)):
                    continue
                p, _ = probability(sup, direction, cat)
                cands.append((p, sup))
            cands.sort(key=lambda t: -t[0])
            assign_cache[key] = cands[:14]
        return assign_cache[key][:n]

    # ── файл 1: оценка лота ──────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Оценка лота"
    cols = [("Лист", 11), ("Производитель", 15), ("Оборудование", 18), ("Наименование", 46),
            ("PN", 17), ("Кол-во", 8), ("ЕИ", 6), ("Категория", 20),
            ("Оценка за ед., USD (от)", 12), ("Оценка за ед., USD (до)", 12),
            ("Сумма, USD (от)", 13), ("Сумма, USD (до)", 13),
            ("Увер.", 7), ("Основа оценки", 30), ("Источник цены", 30),
            ("Поставщик 1 (вер.)", 26), ("Поставщик 2 (вер.)", 26), ("Поставщик 3 (вер.)", 26)]
    sheet_header(ws, cols)
    priced = 0
    for r in demand:
        direction, cat = dir_of(r), r["cat"]
        pr = prices.get((r["pn"] or "").strip().upper())
        sups = top_sups(direction, cat)
        qty = r["qty"] if isinstance(r["qty"], (int, float)) else 0
        row = ws.max_row + 1
        ws.append([r["sheet"], r["man"], r["model"][:40], r["name"][:180], r["pn"], qty, r["unit"], cat,
                   pr["usd_lo"] if pr else None, pr["usd_hi"] if pr else None, None, None,
                   pr["conf"] if pr else "—",
                   (pr.get("basis") or "")[:60] if pr else "оценка в работе (волна цен)",
                   (pr.get("url") or "")[:80] if pr else "",
                   *[f"{s[1]['name'][:22]} ({s[0]}%)" for s in sups] + [""] * (3 - len(sups))])
        if pr:
            priced += 1
            ws.cell(row=row, column=11).value = f"=F{row}*I{row}"
            ws.cell(row=row, column=12).value = f"=F{row}*J{row}"
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=row, column=c)
            cell.font, cell.border = TD, BORD
        if pr:
            ws.cell(row=row, column=13).fill = CONF_FILL.get(pr["conf"], CONF_FILL["C"])
        for c in (9, 10, 11, 12):
            ws.cell(row=row, column=c).number_format = "#,##0.00"
    # итоги
    last = ws.max_row
    ws.append(["ИТОГО (только оценённые)", "", "", "", "", "", "", "", "", "",
               f"=SUM(K2:K{last})", f"=SUM(L2:L{last})", "", f"оценено позиций: {priced} из {len(demand)}", "", "", "", ""])
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=ws.max_row, column=c)
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.border = BORD
    for c in (11, 12):
        ws.cell(row=ws.max_row, column=c).number_format = "#,##0"

    legend = wb.create_sheet("Методика")
    legend.column_dimensions["A"].width = 110
    for t in [
        "Как читать оценку:",
        "Увер. A — найдена цена именно этого PN (розница/дистрибьютор/сток, ссылка в «Источник цены»).",
        "Увер. B — цена близкого аналога той же серии/бренда; вилка отражает разброс каналов (Азия=от, розница ЕС/США=до).",
        "Увер. C — экспертная вилка по типу детали; уточняется запросом поставщику.",
        "Сумма = кол-во × вилка за единицу (формулы, пересчитываются при правке).",
        "Поставщики: топ-3 по вероятности содержательного ответа (история Bitrix > доказанность П-ранга > профиль категории > наличие прямого контакта).",
        "Полный план рассылки с отметками сорсеров: kvant-zip.pages.dev/gt/rfq.html",
        "Цены НЕ включают доставку/посредника/таможню — это рыночная база для торга. Источник: открытые площадки, дата сбора в имени файла.",
    ]:
        legend.append([t])
        legend.cell(row=legend.max_row, column=1).font = TD if legend.max_row > 1 else Font(name="Arial", bold=True, size=11)
        legend.cell(row=legend.max_row, column=1).alignment = Alignment(wrap_text=True)

    out1 = ROOT / "deliverables"
    out1.mkdir(exist_ok=True)
    f1 = out1 / "lot_ocenka.xlsx"
    wb.save(f1)

    # ── файл 2: поставщики по группам ────────────────────────────────────────
    wb2 = Workbook()
    wb2.remove(wb2.active)
    dirs = sorted({dir_of(r) for r in demand})
    cols2 = [("Поставщик", 34), ("Страна", 16), ("Вероятн. ответа", 10), ("Почему (слагаемые)", 44),
             ("Email", 30), ("Телефон", 20), ("Контактное лицо", 22), ("Сайт", 30),
             ("Зацепка для письма", 60), ("Категории пачек", 34)]
    for direction in dirs:
        cats = sorted({r["cat"] for r in demand if dir_of(r) == direction})
        best = {}
        # собираем поставщиков по всем категориям направления
        for cat in cats:
            for p, s in [ (t[0], t[1]) for t in top_sups(direction, cat, 14) ]:
                e = best.setdefault(s["name"], {"sup": s, "p": p, "why": probability(s, direction, cat)[1], "cats": set()})
                e["p"] = max(e["p"], p)
                e["cats"].add(cat)
        name = re.sub(r'[\\/*?:\[\]]', "-", direction)[:31]
        ws2 = wb2.create_sheet(name)
        sheet_header(ws2, cols2)
        for e in sorted(best.values(), key=lambda x: -x["p"]):
            s = e["sup"]
            ws2.append([s["name"], s["country"], f"{e['p']}%", e["why"][:120], s["email"], s["phone"],
                        s["person"], s["site"], s["hook"][:200], ", ".join(sorted(e["cats"]))[:120]])
            for c in range(1, len(cols2) + 1):
                cell = ws2.cell(row=ws2.max_row, column=c)
                cell.font, cell.border = TD, BORD
                if c in (4, 9, 10):
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
    f2 = out1 / "lot_postavshiki.xlsx"
    wb2.save(f2)
    print(f"OK → {f1} и {f2}; оценено {priced}/{len(demand)}")


if __name__ == "__main__":
    main()
