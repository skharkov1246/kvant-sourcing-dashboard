"""CI-зонд v7: экономика проекта по последним 20 сделкам реализации (№895–914).

1) Bitrix: сделки кат.0, порядковые номера 895..914, поле-ссылка на экономику (clck.ru).
2) Разворачиваем clck.ru без выполнения редиректа, классифицируем конечный URL.
3) Пробуем все маршруты доступа токеном аккаунта:
   • личный диск (путь как есть),
   • классические общие папки (могут монтироваться в личный диск — плоский листинг файлов),
   • публичные ключи (если ссылка публичная),
   • общий диск Я360 (vd — ожидаемо недоступен, контрольный выстрел).
4) Всё, что скачалось, парсим openpyxl: листы, ключевые ячейки (закупка/маржа).
Секреты в лог не выводятся.
"""
from __future__ import annotations

import io
import os
import re

import requests

ECON_LINK = "UF_CRM_1740133235324"
DAPI = "https://cloud-api.yandex.net/v1/disk"
SEQ_HI = 914
SEQ_LO = SEQ_HI - 19   # последние 20: 895..914


def bx(method: str, params: dict | None = None) -> dict:
    base = os.environ["BITRIX_WEBHOOK_URL"].rstrip("/")
    r = requests.post(f"{base}/{method}.json", json=params or {}, timeout=60)
    r.raise_for_status()
    return r.json()


def regno(t) -> int:
    m = re.match(r"\s*(\d{1,4})(?:/\d+)?\.", str(t or ""))
    return int(m.group(1)) if m else 0


def cat0_deals() -> list[dict]:
    out, start = [], 0
    while True:
        j = bx("crm.deal.list", {"filter": {"CATEGORY_ID": 0},
                                 "select": ["ID", "TITLE", ECON_LINK], "start": start})
        out += j.get("result", [])
        if "next" not in j:
            return out
        start = j["next"]


def resolve_clck(url: str) -> str:
    """clck.ru/XXX → конечный URL (без выполнения редиректа дальше 1 шага за раз, до 5)."""
    cur = url
    for _ in range(5):
        try:
            r = requests.get(cur, allow_redirects=False, timeout=20)
        except Exception as e:
            return f"EXC:{type(e).__name__}"
        loc = r.headers.get("Location")
        if not loc:
            return cur
        cur = loc if loc.startswith("http") else "https://disk.yandex.ru" + loc
        if "clck.ru" not in cur:
            return cur
    return cur


def classify(url: str) -> tuple[str, str]:
    """→ (класс, полезный путь/ключ)."""
    if url.startswith("EXC:") or not url.startswith("http"):
        return "err", url
    if "/edit/disk/" in url or "/client/disk" in url or "/disk/" in url and "yandex" in url:
        m = re.search(r"/(?:edit/)?disk/([^?]+)", url)
        if m:
            path = requests.utils.unquote(requests.utils.unquote(m.group(1)))
            if path.startswith("vd/") or path.startswith("vd%2F"):
                return "vd", path
            return "disk", path
    if re.search(r"yadi\.sk/[di]/|disk\.yandex\.\w+/[di]/", url):
        return "public", url.split("?")[0]
    return "other", url


def main() -> int:
    tok = (os.getenv("YADISK_TOKEN") or "").strip()
    H = {"Authorization": f"OAuth {tok}"} if tok else {}

    print(f"=== 1. Bitrix: сделки кат.0 c №{SEQ_LO}–{SEQ_HI} и их ссылки на экономику ===")
    deals = cat0_deals()
    sel = []
    for d in deals:
        n = regno(d.get("TITLE"))
        if SEQ_LO <= n <= SEQ_HI:
            sel.append((n, str(d["ID"]), str(d.get("TITLE") or "")[:60],
                        str(d.get(ECON_LINK) or "").strip()))
    sel.sort(key=lambda x: -x[0])
    print(f"всего кат.0: {len(deals)}, в диапазоне: {len(sel)}, со ссылкой: "
          f"{sum(1 for s in sel if s[3])}")

    print("\n=== 2. Разворачиваем clck и классифицируем ===")
    resolved = []
    for n, did, title, link in sel:
        if not link:
            print(f"  №{n} (deal {did}) — ссылки нет · {title}")
            continue
        final = resolve_clck(link)
        cls, path = classify(final)
        resolved.append((n, did, cls, path))
        print(f"  №{n} (deal {did}) → [{cls}] {path[:120]}")

    if not tok:
        print("\nнет YADISK_TOKEN — доступ не проверяем")
        return 0

    print("\n=== 3. Пробуем доступ токеном ===")
    got: list[tuple[int, str, bytes]] = []
    for n, did, cls, path in resolved:
        if cls == "public":
            r = requests.get(DAPI + "/public/resources/download",
                             params={"public_key": path}, headers=H, timeout=30)
            print(f"  №{n} public → {r.status_code} {r.text[:100]}")
            if r.ok:
                href = r.json().get("href")
                f = requests.get(href, timeout=60)
                if f.ok:
                    got.append((n, path, f.content))
        elif cls == "disk":
            p = "disk:/" + path.lstrip("/")
            r = requests.get(DAPI + "/resources/download", params={"path": p}, headers=H, timeout=30)
            print(f"  №{n} disk path={p[:80]!r} → {r.status_code} {r.text[:100]}")
            if r.ok:
                f = requests.get(r.json().get("href"), timeout=60)
                if f.ok:
                    got.append((n, p, f.content))
        elif cls == "vd":
            r = requests.get(DAPI + "/resources/download", params={"path": "disk:/" + path},
                             headers=H, timeout=30)
            print(f"  №{n} vd (контрольный) → {r.status_code} {r.json().get('error', '')[:60] if r.headers.get('content-type','').startswith('application/json') else r.text[:60]}")

    print("\n=== 4. Классические общие папки в личном диске? (плоский листинг xlsx) ===")
    r = requests.get(DAPI + "/resources/files",
                     params={"limit": 200, "media_type": "spreadsheet",
                             "fields": "items.name,items.path,items.size"}, headers=H, timeout=30)
    if r.ok:
        items = r.json().get("items", [])
        econ = [i for i in items if re.match(r"\d{3}_", i.get("name", ""))]
        print(f"  всего таблиц в личном диске: {len(items)}, похожих на экономики (NNN_): {len(econ)}")
        for i in econ[:25]:
            print(f"    {i.get('path')}  ({i.get('size', 0)} B)")
        # если экономики есть в личном диске — качаем нужные номера
        want = {n for n, *_ in resolved} | set(range(SEQ_LO, SEQ_HI + 1))
        for i in econ:
            m = re.match(r"(\d{3,4})_", i.get("name", ""))
            if m and int(m.group(1)) in want and len(got) < 25:
                rr = requests.get(DAPI + "/resources/download", params={"path": i["path"]},
                                  headers=H, timeout=30)
                if rr.ok:
                    f = requests.get(rr.json().get("href"), timeout=60)
                    if f.ok:
                        got.append((int(m.group(1)), i["path"], f.content))
                        print(f"    ✓ скачан {i.get('name')}")
    else:
        print(f"  /resources/files → {r.status_code} {r.text[:120]}")

    print("\n=== 5. Корень личного диска (что смонтировано) ===")
    r = requests.get(DAPI + "/resources", params={"path": "disk:/", "limit": 100,
                     "fields": "_embedded.items.name,_embedded.items.type"}, headers=H, timeout=30)
    if r.ok:
        for i in r.json().get("_embedded", {}).get("items", []):
            print(f"  [{i.get('type')}] {i.get('name')}")
    else:
        print(f"  → {r.status_code} {r.text[:120]}")

    print(f"\n=== 6. Скачано файлов: {len(got)} — парсим openpyxl ===")
    try:
        import openpyxl
    except ImportError:
        openpyxl = None
        print("  openpyxl не установлен")
    for n, src, blob in got if openpyxl else []:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
            print(f"\n  --- №{n} · {str(src)[:70]} · листы: {wb.sheetnames}")
            ws = wb[wb.sheetnames[0]]
            for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=40, max_col=8,
                                                  values_only=True), 1):
                cells = [str(c)[:18] for c in row if c is not None]
                if cells:
                    print(f"    r{ri}: " + " | ".join(cells))
        except Exception as e:
            print(f"  №{n}: ошибка парсинга {type(e).__name__}: {e}")

    print("\n✓ зонд v7 завершён")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
