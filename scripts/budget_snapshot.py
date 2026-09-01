#!/usr/bin/env python3
"""Снимок плановой экономики из «Бюджетов сделок» (Google Drive) → data/budget_snapshot.json.

Источник: папка Google Drive с бюджетами сделок (владелец: 1FUTO4aDl8aK8anw2yxvHnPlxxzObJ0tC).
Каждый файл — книга с листом «Бюджет», где в колонке A подписи, в C/D/E значения
(Сумма с НДС / НДС / Сумма без НДС), суммы в РУБЛЯХ. В шапке листа — точное название
сделки и ссылка на неё в Битриксе → deal_id связывает бюджет со сделкой без угадывания.

Прямого API-доступа к Drive из GitHub Actions нет, поэтому работаем снимком в репозитории
(как data/chat_snapshot.json): агент с доступом к Drive выгружает книги в каталог и запускает

    python scripts/budget_snapshot.py <каталог с *.xlsx> [--filelist list.tsv] [-o data/budget_snapshot.json]

list.tsv (опционально): drive_id \t fileSize \t mimeType \t title [\t modifiedTime] — метаданные
листинга папки; имя книги должно быть <drive_id>.xlsx, тогда к записи добавятся drive_title
и modified (по modified разрешаются дубли, когда на одну сделку заведено несколько книг).

Ограничение: старые книги (сделки примерно до №550, 2024 год) ведены в другом формате —
журнал проводок без блока «Показатели»; из них извлекается только шапка, плановых метрик
не будет. Все такие сделки закрыты, на KPI по открытым это не влияет.
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import pathlib
import re
import sys
import warnings

warnings.filterwarnings("ignore")

# подпись строки (нормализованная) -> ключ снимка; значения из C (с НДС) и E (без НДС)
ROW_LABELS = {
    "выручка": "revenue",
    "итого прямые переменные затраты": "direct_costs",
    "маржинальный доход": "margin_income",
    "% мд": "margin_pct",
    "норма мд": "margin_norm",
    "итого прочие прямые расходы": "other_costs",
    "операционная прибыль": "op_profit",
    "рентабельность оп": "op_margin_pct",
    "прибыль по сделке": "deal_profit",
    "рентабельность по прибыли": "profit_pct",
    "норма рентабельности": "profit_norm",
    "закуп": "purchase",
    "международная логистика": "intl_logistics",
    "курс": "fx_rate",
}
PCT_KEYS = {"margin_pct", "margin_norm", "op_margin_pct", "profit_pct", "profit_norm"}

HEAD_LABELS = {
    "название сделки (точно как в битриксе)": "deal_name",
    "ссылка на сделку в битриксе": "deal_url",
    "заказчик": "customer",
    "исполнитель": "executor",
    "статус сделки": "status",
}


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def _num(v):
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return round(float(v), 4)
    return None


def extract(path: str | pathlib.Path) -> dict:
    """Одна книга бюджета → словарь показателей (или {'error': ...})."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:                                    # noqa: BLE001 — битые книги бывают
        return {"error": f"open: {e}"}
    ws = next((wb[n] for n in wb.sheetnames if _norm(n) == "бюджет"), None)
    if ws is None:
        ws = next((wb[n] for n in wb.sheetnames if "бюджет" in _norm(n)), None)
    if ws is None:
        return {"error": f"no budget sheet: {wb.sheetnames}"}
    out: dict = {}
    try:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 120, 120), max_col=5):
            a = row[0].value if row else None
            if not isinstance(a, str):
                continue
            key = _norm(a)
            # по индексам, не по column_letter: в read_only-режиме у EmptyCell его нет
            vals = [c.value for c in row]                 # 0=A, 1=B, 2=C, 3=D, 4=E
            C = vals[2] if len(vals) > 2 else None
            E = vals[4] if len(vals) > 4 else None
            if key in HEAD_LABELS and C is not None:
                out.setdefault(HEAD_LABELS[key], str(C).strip())
            if key in ROW_LABELS:
                k = ROW_LABELS[key]
                if k == "fx_rate":
                    out.setdefault(k, _num(C))
                elif k in PCT_KEYS:
                    v = _num(E)
                    out[k] = v if v is not None else _num(C)
                elif out.get(k) is None:
                    out[k] = _num(C)                       # с НДС
                    out[k + "_net"] = _num(E)              # без НДС
    except Exception as e:                                    # noqa: BLE001
        return {**out, "error": f"parse: {e}"}
    m = re.search(r"/deal/details/(\d+)", out.get("deal_url") or "")
    if m:
        out["deal_id"] = int(m.group(1))
    return out


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("dir", help="каталог с выгруженными книгами бюджетов (*.xlsx)")
    ap.add_argument("--filelist", help="TSV листинга папки Drive: id, size, mime, title")
    ap.add_argument("-o", "--out", default="data/budget_snapshot.json")
    args = ap.parse_args()

    meta: dict[str, tuple[str, str]] = {}
    if args.filelist:
        for line in pathlib.Path(args.filelist).read_text().splitlines():
            parts = line.split("\t")
            if len(parts) >= 4:
                meta[parts[0]] = (parts[3], parts[4][:10] if len(parts) > 4 else "")

    budgets, errors = [], []
    files = sorted(pathlib.Path(args.dir).glob("*.xlsx"))
    for f in files:
        rec = extract(f)
        rec["drive_id"] = f.stem
        if f.stem in meta:
            rec["drive_title"], rec["modified"] = meta[f.stem]
        (errors if rec.get("error") else budgets).append(rec)

    snap = {
        "generatedAt": dt.date.today().isoformat(),
        "folder": "1FUTO4aDl8aK8anw2yxvHnPlxxzObJ0tC",
        "nFiles": len(files), "nParsed": len(budgets), "nErrors": len(errors),
        "budgets": budgets, "errors": errors,
    }
    outp = pathlib.Path(args.out)
    outp.parent.mkdir(parents=True, exist_ok=True)
    outp.write_text(json.dumps(snap, ensure_ascii=False, indent=None, separators=(",", ":")) + "\n")
    with_id = sum(1 for b in budgets if b.get("deal_id"))
    print(f"{outp}: книг {len(files)}, разобрано {len(budgets)} (с deal_id {with_id}), ошибок {len(errors)}",
          file=sys.stderr)


if __name__ == "__main__":
    main()
